# -*- coding: utf-8 -*-
"""
This is the final version of score-report script

Advanced logging and configuration options are included

Offline mini-server is added
"""

from smtplib import SMTP_SSL, SMTPException
from time import sleep

from secret import *    # MAILBOX & PASSWD

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook  # read xlsx
from scipy.stats import rankdata  # calculate ranking

# Log file
import logging

# My address
me = MAILBOX + '@163.com'

# Log file
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

    # Handler
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter('%(asctime)s - %(message)s'))
fh = logging.FileHandler('send_score_report.log')
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))

    # add to logger
logger.addHandler(ch)
logger.addHandler(fh)

# configure
import configparser
_config = configparser.ConfigParser()
_config.read('mailsend.ini')

_debugmode = _config.getboolean('Debug', 'debugmode')
_offlinemode = _config.getboolean('Debug', 'offlinemode')
_offlinedumpfile = _config['Debug']['offlinedumpfile']
_offlineerrorpercentile = _config.getfloat('Debug', 'offlineerrorpercentile')
_dumpaddr = _config['Debug']['dumpaddress']
_trials = _config.getint('Debug','trials')

_starting_no = _config.getint('Sending','starting_no')
_addr_char = _config['Sending']['addr_char']
_continue_unsent = _config.getboolean('Sending', 'continue_unsent')
_waitsec = _config.getfloat('Sending', 'waitsec')

_contents = _config['Mail']['contents']
_sender = _config['Mail']['sender']
_subject = _config['Mail']['subject']
_score_file = _config['Data']['filename']
_score_range = _config['Data']['score_range']
_data_size = _config.getint('Data', 'size')


def writemail(me, no, stuid, addr, score, rank, chname):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = _subject + '{}'.format(stuid)
    msg['From'] = _sender + '<{}>'.format(me)
    msg['To'] = _dumpaddr if _debugmode else addr
    html = _contents.format(name=chname, score=score, ranking=rank)
    msg.attach(MIMEText(html, 'html', 'utf8'))

    return msg


def iter_testers():
    # Load Workbook
    wb = load_workbook(filename=_score_file, read_only=True)
    ws = wb.active

    # Calculate rankings
    scores = [_[0].value if _[0].value is not None else 0 for _ in ws[_score_range]]
    rankings = [int(_data_size + 1 - _) for _ in rankdata(scores,method='max')]

    # iterate through testers
    for i, line in enumerate(ws.iter_rows()):
        if i >= 1: # clear header
            no, stuid, engname, addr, score, chname = map(lambda x:getattr(x, 'value'), line)
            ranking = rankings[i-1]
            yield (no, stuid, addr, score, ranking, chname)


def send_condition(no, stuid, addr): #when need a retrial, start from a new number(included)
    if _debugmode:
        return int(no) >= _starting_no and _addr_char in addr and no <= _trials
    else:
        return int(no) >= _starting_no and _addr_char in addr


# An offline mail server for debugging
import random

class _OfflineMailServer:
    def __init__(self):
        with open(_offlinedumpfile,'w'):pass
        random.seed()

    def login(self, mailbox, passwd):
        pass

    def send_message(self, msg):
        if random.randint(0,100) < _offlineerrorpercentile:
            raise SMTPException
        else:
            with open(_offlinedumpfile, 'a') as fout:
                fout.write('\n'+'='*30+'\n')
                fout.write('Subject:{}\n'.format(msg['subject']))
                fout.write('From:{}\n'.format(msg['from']))
                fout.write('To:{}\n'.format(msg['to']))
                fout.write(msg.get_payload(0).get_payload(decode=True).decode('utf-8'))
                fout.write('\n'*3)

    def quit(self):
        pass

# My address
me = MAILBOX + '@163.com'

try:
    # Server
    s = SMTP_SSL('smtp.163.com',465) if not _offlinemode else _OfflineMailServer()
    s.login(MAILBOX, PASSWD)
    logger.info('login successful.')

    for no, stuid, addr, score, ranking, chname in iter_testers():
        if send_condition(no, stuid, addr):
            msg = writemail(me, no, stuid, addr, score, ranking, chname)
            while True:
                try:
                    s.send_message(msg)
                    logger.info('no.{} is sent!'.format(no))
                    break
                except SMTPException as e:
                    logger.info('SMTP Exception occurred when processing no.{0}: {1}'.format(no, stuid))
                    logger.debug(e.strerror)

                    s.quit()
                    logger.info('Try again in {} seconds...'.format(_waitsec))
                    sleep(_waitsec)
                    s = SMTP_SSL('smtp.163.com',465) if not _offlinemode else s
                    s.login(MAILBOX, PASSWD)
                    logger.info('login successful.')

    else:
        s.quit()
        logger.info('All mails succeesfully sent.')
        if not _debugmode and _continue_unsent:
            _config['Sending']['starting_no'] = str(no + 1)
            with open('mailsend.ini', 'w') as f:
                _config.write(f)

except: # in case of unexpected error
    if not _debugmode and _continue_unsent:
        logger.info('Unexpected Exception')
        _config['Sending']['starting_no'] = str(no)
        with open('mailsend.ini', 'w') as f:
            _config.write(f)
    raise